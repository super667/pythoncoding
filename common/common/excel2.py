import xlwt
import openpyxl
import os


class CExcel_openpyxl:
    def __init__(self, filepath):
        self.filepath = os.path.abspath(filepath)

    def open(self):
        import openpyxl
        self.xlsx = openpyxl.load_workbook(self.filepath)

    def save(self, filename_xlsx):
        self.xlsx.save(filename_xlsx)

    def save(self):
        self.xlsx.save(self.filepath)

    def close(self):
        # self.xlBook.Close(False)
        pass

    def getSheetCount(self):
        return len(self.xlsx.get_sheet_names())

    def getSheetNameList(self):
        sheet_name_list = self.xlsx.get_sheet_names()
        return sheet_name_list

    def getRowCount(self, sheet_name):
        # sheet = self.xlBook.Worksheets(sheet_name)
        # return sheet.UsedRange.Rows.Count
        pass

    def getColumnCount(self, sheet_name):
        # sheet = self.xlBook.Worksheets(sheet_name)
        # return sheet.UsedRange.Columns.Count
        sheet = self.xlsx.get_sheet_by_name(sheet_name)
        return sheet.max_row - sheet.min_row

    def getDimensions(self, sheet_name):
        sheet = self.xlsx.get_sheet_by_name(sheet_name)
        return (sheet.dimensions, sheet.min_row, sheet.max_row, sheet.min_column, sheet.max_column)

    def getCellValue(self, sheet_name, ROW, COL):
        sheet = self.xlsx.get_sheet_by_name(sheet_name)
        return sheet.cell(row=ROW, column=COL).value


    def setCellValue(self, sheet_name, row, col, value):
        sheet = self.xlsx.get_sheet_by_name(sheet_name)
        sheet.cell(row, col, value=value)

    def setCellRed(self, sheet_name, row, col):
        # sheet = self.xlBook.Worksheets(sheet_name)
        # sheet.Cells(row, col).Interior.ColorIndex = 3
        pass

    def deleteRowValue(self, sheet_name, row, num):
        sheet = self.xlsx.get_sheet_by_name(sheet_name)
        sheet.delete_rows(row, num)

